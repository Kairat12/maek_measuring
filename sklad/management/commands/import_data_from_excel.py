import os
from django.core.management.base import BaseCommand
from sklad.models import MainSklad  # Замените на вашу модель
from openpyxl import load_workbook
from datetime import datetime

class Command(BaseCommand):
    help = 'Import data from Excel'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR('File does not exist'))
            return

        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            data_count = 0
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index < 4:
                    continue
                data_count = data_count + 1
                sklad, created = MainSklad.objects.get_or_create(
                    item_number=row[0] if row[0] else '',
                    name=row[1] if row[1] else '',
                    quantity=row[3] if row[3] else 0,
                    price=row[4] if row[4] else 0,
                    sum=row[5] if row[4] else 0,
                    number_sklad=row[7] if row[7] else '',
                    defaults={'unit': row[2] if row[2] else '',
                              'date_receipt': datetime.strptime(str(row[6]), '%d.%m.%y').strftime('%Y-%m-%d')
                              if row[6] else datetime.strptime('1970-01-01', '%Y-%m-%d').strftime('%Y-%m-%d'),
                              'responsible': row[8] if row[8] else '',
                              'contractor': row[9] if row[9] else '',
                              'agreement': row[10] if row[10] else '',
                              }
                )
                if created:
                    self.stdout.write(f"Создана запись: {sklad}")
                else:
                    sklad.item_number = row[0] if row[0] else ''
                    sklad.name = row[1] if row[1] else ''
                    sklad.unit = row[2] if row[2] else ''
                    sklad.quantity = row[3] if row[3] else 0
                    sklad.price = row[4] if row[4] else 0
                    sklad.sum = row[5] if row[5] else 0
                    print(f'{row[6]} - {data_count}')
                    sklad.date_receipt = datetime.strptime(str(row[6]), '%d.%m.%y').strftime('%Y-%m-%d') if row[6] is not None else datetime.strptime('1970-01-01', '%Y-%m-%d').strftime('%Y-%m-%d')
                    sklad.number_sklad = row[7] if row[7] else ''
                    sklad.responsible = row[8] if row[8] else ''
                    sklad.contractor = row[9] if row[6] else ''
                    sklad.agreement = row[10] if row[6] else ''
                    sklad.save()
                    self.stdout.write(f"Обновлена запись: {sklad}")

            # self.stdout.write(self.style.SUCCESS('Data imported successfully'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'An error occurred: {e}'))
